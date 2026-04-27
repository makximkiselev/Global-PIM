# backend/app/api/routes/templates.py
from __future__ import annotations

from copy import deepcopy
import io
from datetime import datetime, timezone
import time
from pathlib import Path
from typing import Any, Dict, List

from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel, Field
from fastapi.responses import StreamingResponse

from app.core.json_store import read_doc, with_lock
from app.storage.json_store import (
    ensure_global_attribute,
    load_dictionaries_db,
    load_templates_db,
    save_templates_db,
    new_id,
    slugify_code,
)
from app.storage.relational_pim_store import (
    bulk_upsert_product_items,
    load_catalog_nodes,
    load_category_mappings,
    load_template_editor_payload,
    query_products_full,
)
from app.core.master_templates import (
    base_field_by_code,
    base_field_by_name,
    base_template_fields,
    is_deprecated_template_code,
    is_deprecated_template_name,
    split_template_attrs,
)

router = APIRouter(prefix="/templates", tags=["templates"])
_DEFAULT_ATTRS_CACHE: List[Dict[str, Any]] | None = None

BASE_DIR = Path(__file__).resolve().parents[3]
DATA_DIR = BASE_DIR / "data"
MARKETPLACES_DIR = DATA_DIR / "marketplaces"
CATEGORY_MAPPING_PATH = MARKETPLACES_DIR / "category_mapping.json"
YANDEX_CATEGORY_PARAMS_PATH = MARKETPLACES_DIR / "yandex_market" / "category_parameters.json"
YANDEX_CATEGORIES_TREE_PATH = MARKETPLACES_DIR / "yandex_market" / "categories_tree.json"
OZON_CATEGORY_ATTRS_PATH = MARKETPLACES_DIR / "ozon" / "category_attributes.json"
OZON_CATEGORIES_TREE_PATH = MARKETPLACES_DIR / "ozon" / "categories_tree.json"
CATALOG_NODES_PATH = DATA_DIR / "catalog_nodes.json"
_EDITOR_REFERENCE_CACHE_TTL_SECONDS = 300.0
_editor_reference_cache: Dict[str, Any] = {"ts": 0.0, "payload": None}


# =========================
# helpers
# =========================
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


ALLOWED_TYPES = {"text", "number", "select", "bool", "date", "json"}
ALLOWED_SCOPES = {"common", "variant"}

TYPE_RU: Dict[str, str] = {
    "text": "Текст",
    "number": "Число",
    "select": "Список",
    "bool": "Да/Нет",
    "date": "Дата",
    "json": "JSON",
}
TYPE_EN_BY_RU = {v.lower(): k for k, v in TYPE_RU.items()}

SCOPE_RU: Dict[str, str] = {
    "common": "Общий",
    "variant": "Вариант",
}
SCOPE_EN_BY_RU = {v.lower(): k for k, v in SCOPE_RU.items()}

def _boolish(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "+", "да", "д", "ok", "✅"}


def _norm_type(v: Any) -> str:
    s = (str(v).strip().lower() if v is not None else "") or "text"
    if s in TYPE_EN_BY_RU:
        s = TYPE_EN_BY_RU[s]
    if s not in ALLOWED_TYPES:
        s = "text"
    return s


def _norm_scope(v: Any) -> str:
    s = (str(v).strip().lower() if v is not None else "") or "common"
    if s in SCOPE_EN_BY_RU:
        s = SCOPE_EN_BY_RU[s]
    if s not in ALLOWED_SCOPES:
        s = "common"
    return s


def _build_default_attrs() -> List[Dict[str, Any]]:
    global _DEFAULT_ATTRS_CACHE
    if _DEFAULT_ATTRS_CACHE is not None:
        return deepcopy(_DEFAULT_ATTRS_CACHE)

    out: List[Dict[str, Any]] = []
    for idx, a in enumerate(base_template_fields()):
        global_attr = ensure_global_attribute(
            title=str(a["name"]),
            type_=str(a["type"]),
            code=str(a["code"]),
            scope="variant" if str(a.get("scope") or "") == "variant" else "feature",
        )
        options = {
            "dict_id": str(global_attr.get("dict_id") or "").strip() or None,
            "attribute_id": str(global_attr.get("id") or "").strip() or None,
            "param_group": str(a.get("param_group") or "").strip() or None,
            "layer": "base",
            "system_key": str(a.get("key") or "").strip() or None,
        }
        out.append(
            {
                "id": new_id(),
                "name": a["name"],
                "code": a["code"],
                "type": _norm_type(a["type"]),
                "required": bool(a.get("required")),
                "scope": _norm_scope(a.get("scope")),
                "attribute_id": str(global_attr.get("id") or "").strip() or None,
                "options": options,
                "position": idx,
                "locked": True,
            }
        )
    _DEFAULT_ATTRS_CACHE = out
    return deepcopy(out)


def _load_catalog_nodes() -> List[Dict[str, Any]]:
    return load_catalog_nodes()


def _load_products_doc() -> Dict[str, Any]:
    return {"version": 1, "items": query_products_full()}


def _catalog_path(nodes: List[Dict[str, Any]], category_id: str) -> List[Dict[str, str]]:
    by_id = {str(n.get("id") or ""): n for n in nodes if isinstance(n, dict)}
    chain: List[Dict[str, str]] = []
    seen: set[str] = set()
    cur = by_id.get(str(category_id or "").strip())
    while cur:
        cid = str(cur.get("id") or "").strip()
        if not cid or cid in seen:
            break
        seen.add(cid)
        chain.append({"id": cid, "name": str(cur.get("name") or cid)})
        pid = str(cur.get("parent_id") or "").strip()
        cur = by_id.get(pid) if pid else None
    chain.reverse()
    return chain


def _feature_skeleton_attrs(attrs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for attr in attrs or []:
        if not isinstance(attr, dict):
            continue
        code = str(attr.get("code") or "").strip()
        if not code or code in seen:
            continue
        options = attr.get("options") if isinstance(attr.get("options"), dict) else {}
        layer = str(options.get("layer") or "").strip().lower()
        group = str(options.get("param_group") or "").strip()
        if layer == "base" or group in {"Описание", "Медиа"}:
            continue
        seen.add(code)
        out.append({"code": code, "name": str(attr.get("name") or code)})
    return out


def _descendant_category_ids(nodes: List[Dict[str, Any]], root_category_id: str) -> List[str]:
    children: Dict[str, List[str]] = {}
    for node in nodes or []:
        if not isinstance(node, dict):
            continue
        pid = str(node.get("parent_id") or "").strip()
        nid = str(node.get("id") or "").strip()
        if nid:
            children.setdefault(pid, []).append(nid)
    out: List[str] = []
    stack = [str(root_category_id or "").strip()]
    seen: set[str] = set()
    while stack:
        cid = stack.pop()
        if not cid or cid in seen:
            continue
        seen.add(cid)
        out.append(cid)
        stack.extend(children.get(cid, []))
    return out


def _merge_feature_skeleton(existing_features: Any, skeleton: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    existing = existing_features if isinstance(existing_features, list) else []
    existing_by_code: Dict[str, Dict[str, Any]] = {}
    ordered_extra: List[Dict[str, Any]] = []
    for item in existing:
        if not isinstance(item, dict):
            continue
        code = str(item.get("code") or "").strip()
        if code and code not in existing_by_code:
            existing_by_code[code] = item
        else:
            ordered_extra.append(item)
    merged: List[Dict[str, Any]] = []
    used_codes: set[str] = set()
    for sk in skeleton:
        code = str(sk.get("code") or "").strip()
        if not code:
            continue
        used_codes.add(code)
        cur = existing_by_code.get(code)
        if cur:
            merged.append(
                {
                    "code": code,
                    "name": str(cur.get("name") or sk.get("name") or code),
                    "restore": str(cur.get("restore") or ""),
                    "store77": str(cur.get("store77") or ""),
                    "selected": str(cur.get("selected") or "custom"),
                    "value": str(cur.get("value") or ""),
                }
            )
            continue
        merged.append(
            {
                "code": code,
                "name": str(sk.get("name") or code),
                "restore": "",
                "store77": "",
                "selected": "custom",
                "value": "",
            }
        )
    for item in existing:
        if not isinstance(item, dict):
            continue
        code = str(item.get("code") or "").strip()
        if code and code in used_codes:
            continue
        merged.append(item)
    merged.extend(ordered_extra)
    return merged


class ApplyTemplateToProductsReq(BaseModel):
    include_descendants: bool = True
    dry_run: bool = False


def _ensure_default_attrs(attrs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    attrs = [
        a
        for a in (attrs or [])
        if isinstance(a, dict)
        and not is_deprecated_template_code(a.get("code"))
        and not is_deprecated_template_name(a.get("name"))
    ]
    by_code: Dict[str, Dict[str, Any]] = {}
    by_name: Dict[str, Dict[str, Any]] = {}
    for a in attrs:
        code = (a.get("code") or "").strip().lower()
        if code and code not in by_code:
            by_code[code] = a
        name = (a.get("name") or "").strip().lower()
        if name and name not in by_name:
            by_name[name] = a

    for d in base_template_fields():
        code = d["code"].strip().lower()
        name = str(d["name"]).strip().lower()
        global_attr = ensure_global_attribute(
            title=str(d["name"]),
            type_=str(d["type"]),
            code=str(d["code"]),
            scope="variant" if str(d.get("scope") or "") == "variant" else "feature",
        )
        required = bool(d.get("required"))
        base_options = {
            "dict_id": str(global_attr.get("dict_id") or "").strip() or None,
            "attribute_id": str(global_attr.get("id") or "").strip() or None,
            "param_group": str(d.get("param_group") or "").strip() or None,
            "layer": "base",
            "system_key": str(d.get("key") or "").strip() or None,
        }
        existing = by_code.get(code) or by_name.get(name)
        if existing:
            existing["locked"] = True
            existing["code"] = d["code"]
            existing["name"] = d["name"]
            existing["required"] = required
            existing["type"] = _norm_type(d["type"])
            existing["scope"] = _norm_scope(d.get("scope"))
            existing["attribute_id"] = str(global_attr.get("id") or "").strip() or None
            options = existing.get("options") if isinstance(existing.get("options"), dict) else {}
            existing["options"] = {**options, **base_options}
            continue
        attrs.append(
            {
                "id": new_id(),
                "name": d["name"],
                "code": d["code"],
                "type": _norm_type(d["type"]),
                "required": required,
                "scope": _norm_scope(d.get("scope")),
                "attribute_id": str(global_attr.get("id") or "").strip() or None,
                "options": base_options,
                "position": len(attrs),
                "locked": True,
            }
        )

    deduped: List[Dict[str, Any]] = []
    seen_base_codes: set[str] = set()
    for attr in attrs:
        code = str(attr.get("code") or "").strip().lower()
        name = str(attr.get("name") or "").strip()
        base_def = base_field_by_code(code) or base_field_by_name(name)
        if (is_deprecated_template_code(code) or is_deprecated_template_name(name)) and bool(attr.get("locked")):
            continue
        if base_def:
            canonical_code = str(base_def.get("code") or "").strip().lower()
            if canonical_code in seen_base_codes:
                continue
            seen_base_codes.add(canonical_code)
            attr["code"] = str(base_def.get("code") or attr.get("code") or "").strip()
            attr["name"] = str(base_def.get("name") or attr.get("name") or "").strip()
        deduped.append(attr)

    attrs = deduped
    attrs.sort(key=lambda x: int(x.get("position", 0)))
    for i, a in enumerate(attrs):
        a["position"] = i
    return attrs

def _get_catalog_nodes() -> List[Dict[str, Any]]:
    try:
        # Editor bootstrap needs only the category tree. It must not pull
        # product counts for the whole catalog, because that path is much
        # heavier and blocks opening the model workspace.
        return load_catalog_nodes()
    except Exception:
        return []


def _dedupe_codes(attrs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    out = []
    for a in attrs:
        code = (a.get("code") or "").strip().lower()
        if not code:
            code = slugify_code(a.get("name") or "")
        base = code
        j = 2
        while code in seen:
            code = f"{base}_{j}"
            j += 1
        seen.add(code)
        a["code"] = code
        out.append(a)
    return out


def _normalize_attributes(attrs_in: Any) -> List[Dict[str, Any]]:
    if not isinstance(attrs_in, list):
        raise HTTPException(status_code=400, detail="attributes must be a list")

    out: List[Dict[str, Any]] = []
    for idx, a in enumerate(attrs_in):
        if not isinstance(a, dict):
            continue

        name = (a.get("name") or "").strip()
        if not name:
            continue

        code = (a.get("code") or "").strip().lower()
        if not code:
            code = slugify_code(name)

        code_norm = code.strip().lower()
        base_def = base_field_by_code(code_norm)
        options = a.get("options") or {}
        if not isinstance(options, dict):
            options = {}
        if base_def:
            options = {
                **options,
                "layer": "base",
                "system_key": base_def.get("key"),
                "param_group": options.get("param_group") or base_def.get("param_group"),
            }
        else:
            options = {
                **options,
                "layer": options.get("layer") or "category",
            }

        out.append(
            {
                "id": a.get("id") or new_id(),
                "name": name,
                "code": code,
                "type": _norm_type(a.get("type")),
                "required": bool(a.get("required")),
                "scope": _norm_scope(a.get("scope")),
                "options": options,
                "position": int(a.get("position") or idx),
                "locked": bool(a.get("locked") or bool(base_def)),
                "attribute_id": a.get("attribute_id") or options.get("attribute_id"),
            }
        )

    out.sort(key=lambda x: int(x.get("position", 0)))
    for i, a in enumerate(out):
        a["position"] = i

    return _dedupe_codes(out)


def _templates_by_category(db: Dict[str, Any]) -> Dict[str, List[str]]:
    """
    ✅ Новая логика: много шаблонов на категорию.
    Источник истины: templates[*].category_id.

    Также поддерживаем legacy поля:
    - category_to_template: {category_id: template_id}
    - category_to_templates: {category_id: [template_id,...]}
    """
    out: Dict[str, List[str]] = {}

    templates = db.get("templates") or {}

    # 1) из templates[*].category_id
    if isinstance(templates, dict):
        for tid, t in templates.items():
            if not isinstance(t, dict):
                continue
            cid = (t.get("category_id") or "").strip()
            if not cid:
                continue
            out.setdefault(cid, []).append(str(t.get("id") or tid))

    # 2) из category_to_templates (если есть)
    cat_to_tpls = db.get("category_to_templates") or {}
    if isinstance(cat_to_tpls, dict):
        for cid, v in cat_to_tpls.items():
            if isinstance(v, list):
                for tid in v:
                    if tid:
                        out.setdefault(str(cid), []).append(str(tid))

    # 3) из legacy category_to_template (если есть)
    cat_to_tpl = db.get("category_to_template") or {}
    if isinstance(cat_to_tpl, dict):
        for cid, tid in cat_to_tpl.items():
            if tid:
                out.setdefault(str(cid), []).append(str(tid))

    # dedupe + стабильный порядок (по updated_at desc, затем name)
    def tpl_sort_key(tid: str):
        t = (templates or {}).get(tid) or {}
        upd = t.get("updated_at") or ""
        nm = t.get("name") or ""
        return (upd, nm)

    for cid, arr in out.items():
        uniq = []
        seen = set()
        for tid in arr:
            if tid in seen:
                continue
            seen.add(tid)
            uniq.append(tid)
        uniq.sort(key=tpl_sort_key, reverse=True)  # самые свежие первыми
        out[cid] = uniq

    return out


def _template_master_payload(template: Dict[str, Any], attrs: List[Dict[str, Any]], *, include_sources: bool = True) -> Dict[str, Any]:
    split = split_template_attrs(attrs)
    meta = template.get("meta") if isinstance(template.get("meta"), dict) else {}
    sources = meta.get("sources") if isinstance(meta.get("sources"), dict) else {}
    base_attrs = split["base"]
    category_attrs = split["category"]

    if include_sources and isinstance(template, dict):
        category_id = str(template.get("category_id") or "").strip()
        provider_mapped_rows: Dict[str, int] = {"yandex_market": 0, "ozon": 0}
        for attr in attrs or []:
            if not isinstance(attr, dict):
                continue
            options = attr.get("options") if isinstance(attr.get("options"), dict) else {}
            sb = options.get("source_bindings") if isinstance(options.get("source_bindings"), dict) else {}
            for provider in ("yandex_market", "ozon"):
                p = sb.get(provider) if isinstance(sb.get(provider), dict) else {}
                if str(p.get("id") or "").strip() or str(p.get("name") or "").strip():
                    provider_mapped_rows[provider] += 1

        cat_map = load_category_mappings()
        category_mapping = (cat_map.get(category_id) or {}) if isinstance(cat_map.get(category_id), dict) else {}
        yandex_category_id = str(category_mapping.get("yandex_market") or "").strip()
        ozon_category_id = str(category_mapping.get("ozon") or "").strip()

        def _provider_cat_name(tree_path: Path, provider_category_id: str) -> str:
            lookup_id = str(provider_category_id or "").strip()
            if not lookup_id:
                return ""
            if lookup_id.startswith("type:"):
                parts = lookup_id.split(":")
                if len(parts) >= 3:
                    lookup_id = str(parts[1] or "").strip()
            tree_doc = read_doc(tree_path, default={})
            flat = tree_doc.get("flat") if isinstance(tree_doc, dict) else []
            if not isinstance(flat, list):
                return ""
            for row in flat:
                if not isinstance(row, dict):
                    continue
                if str(row.get("id") or "").strip() != lookup_id:
                    continue
                return str(row.get("path") or row.get("name") or "").strip()
            return ""

        def _yandex_stats(provider_category_id: str) -> Dict[str, int]:
            params_count = 0
            required_params_count = 0
            params_doc = read_doc(YANDEX_CATEGORY_PARAMS_PATH, default={"items": {}})
            params_items = params_doc.get("items") if isinstance(params_doc, dict) else {}
            if isinstance(params_items, dict) and provider_category_id:
                prow = params_items.get(provider_category_id)
                raw = prow.get("raw") if isinstance(prow, dict) and isinstance(prow.get("raw"), dict) else {}
                result = raw.get("result") if isinstance(raw, dict) else {}
                params = result.get("parameters") if isinstance(result, dict) else []
                if isinstance(params, list):
                    params_count = len([x for x in params if isinstance(x, dict)])
                    required_params_count = len([x for x in params if isinstance(x, dict) and bool(x.get("required") or False)])
            return {"params_count": params_count, "required_params_count": required_params_count}

        def _ozon_stats(provider_category_id: str) -> Dict[str, int]:
            lookup_id = str(provider_category_id or "").strip()
            if lookup_id.startswith("type:"):
                parts = lookup_id.split(":")
                if len(parts) >= 3:
                    lookup_id = str(parts[1] or "").strip()
            attrs_count = 0
            required_count = 0
            attrs_doc = read_doc(OZON_CATEGORY_ATTRS_PATH, default={"items": {}})
            attrs_items = attrs_doc.get("items") if isinstance(attrs_doc, dict) else {}
            if isinstance(attrs_items, dict) and lookup_id:
                prow = attrs_items.get(lookup_id)
                attrs = prow.get("attributes") if isinstance(prow, dict) and isinstance(prow.get("attributes"), list) else []
                if isinstance(attrs, list):
                    attrs_count = len([x for x in attrs if isinstance(x, dict)])
                    required_count = len([x for x in attrs if isinstance(x, dict) and bool(x.get("is_required") or x.get("required") or False)])
            return {"params_count": attrs_count, "required_params_count": required_count}

        next_sources: Dict[str, Any] = dict(sources) if isinstance(sources, dict) else {}
        if yandex_category_id or provider_mapped_rows["yandex_market"]:
            y_stats = _yandex_stats(yandex_category_id)
            next_sources["yandex_market"] = {
                "enabled": True,
                "mode": "structure_source",
                "category_id": yandex_category_id or None,
                "category_name": _provider_cat_name(YANDEX_CATEGORIES_TREE_PATH, yandex_category_id) or None,
                "params_count": y_stats["params_count"],
                "required_params_count": y_stats["required_params_count"],
                "mapped_rows": provider_mapped_rows["yandex_market"],
            }
        if ozon_category_id or provider_mapped_rows["ozon"]:
            o_stats = _ozon_stats(ozon_category_id)
            next_sources["ozon"] = {
                "enabled": True,
                "mode": "structure_source",
                "category_id": ozon_category_id or None,
                "category_name": _provider_cat_name(OZON_CATEGORIES_TREE_PATH, ozon_category_id) or None,
                "params_count": o_stats["params_count"],
                "required_params_count": o_stats["required_params_count"],
                "mapped_rows": provider_mapped_rows["ozon"],
            }
        sources = next_sources

    template_meta = template.get("meta") if isinstance(template.get("meta"), dict) else {}
    master_meta = template_meta.get("master_template") if isinstance(template_meta.get("master_template"), dict) else {}
    row_count = int(master_meta.get("row_count") or len(category_attrs) or 0)
    confirmed_count = int(master_meta.get("confirmed_count") or 0)
    if row_count > 0 and confirmed_count >= row_count:
        readiness = "ready"
    elif confirmed_count > 0:
        readiness = "in_progress"
    else:
        readiness = "draft"

    return {
        "version": 2,
        "base_attributes": base_attrs,
        "category_attributes": category_attrs,
        "stats": {
            "base_count": len(base_attrs),
            "category_count": len(category_attrs),
            "required_count": sum(1 for x in attrs if bool(x.get("required"))),
            "total_count": len(attrs),
            "row_count": row_count,
            "confirmed_count": confirmed_count,
        },
        "sources": sources,
        "status": readiness,
    }


# ============================================================
# ✅ FIXED ROUTES FIRST (ВАЖНО: ДО /{template_id})
# ============================================================

# =========================
# List endpoints (для маппинга)
# =========================
@router.get("/list")
def templates_list() -> Dict[str, Any]:
    """
    ✅ Возвращает список мастер-шаблонов (не дерево каталога).
    Это нужно для страницы "Маппинг конкурентов".
    """
    db = load_templates_db()
    templates = db.get("templates") or {}
    items: List[Dict[str, Any]] = []

    if isinstance(templates, dict):
        for tid, t in templates.items():
            if not isinstance(t, dict):
                continue
            items.append(
                {
                    "id": str(t.get("id") or tid),
                    "category_id": t.get("category_id"),
                    "name": t.get("name") or "Без названия",
                    "updated_at": t.get("updated_at"),
                    "created_at": t.get("created_at"),
                    "master": _template_master_payload(
                        t,
                        _ensure_default_attrs((db.get("attributes", {}) or {}).get(str(t.get("id") or tid), []) or []),
                        include_sources=False,
                    ),
                }
            )

    items.sort(key=lambda x: (x.get("name") or "").lower())
    return {"ok": True, "items": items}


# =========================
# Tree endpoints
# =========================
@router.get("/tree")
def templates_tree() -> Dict[str, Any]:
    """
    Возвращает nodes каталога + template_ids на узлах.
    legacy: template_id = первый (для старых мест, где ожидается один).
    """
    db = load_templates_db()
    cat_map = _templates_by_category(db)

    nodes = _get_catalog_nodes()
    for n in nodes:
        cid = str(n.get("id"))
        tids = cat_map.get(cid, []) or []
        n["template_ids"] = tids
        n["template_id"] = tids[0] if tids else None  # legacy

    return {"nodes": nodes}


@router.get("/by-category/{category_id}")
def get_by_category(category_id: str) -> Dict[str, Any]:
    """
    ⚠️ Backward compatible:
    раньше было {template, attributes} для одного шаблона.
    Теперь:
      - templates: список шаблонов категории
      - template/attributes: legacy первый шаблон
    """
    db = load_templates_db()
    templates = db.get("templates", {}) or {}
    attrs_by_tpl = db.get("attributes", {}) or {}

    cat_map = _templates_by_category(db)
    tids = cat_map.get(category_id, []) or []

    tpl_list = []
    for tid in tids:
        t = templates.get(tid)
        if isinstance(t, dict):
            tpl_list.append(t)

    # legacy first
    if tids:
        first_id = tids[0]
        tpl = templates.get(first_id)
        attrs = attrs_by_tpl.get(first_id, []) or []
        attrs = _ensure_default_attrs(attrs)
    else:
        tpl = None
        attrs = []

    master = _template_master_payload(tpl, attrs, include_sources=False) if isinstance(tpl, dict) else {
        "version": 2,
        "base_attributes": _build_default_attrs(),
        "category_attributes": [],
        "stats": {
            "base_count": len(_build_default_attrs()),
            "category_count": 0,
            "required_count": sum(1 for x in _build_default_attrs() if bool(x.get("required"))),
            "total_count": len(_build_default_attrs()),
        },
        "sources": {},
    }

    return {
        "templates": tpl_list,
        "template": tpl,
        "attributes": attrs,
        "base_attributes": master["base_attributes"],
        "category_attributes": master["category_attributes"],
        "master": master,
    }


def _template_editor_reference_payload() -> Dict[str, Any]:
    now = time.time()
    cached = _editor_reference_cache.get("payload")
    cached_ts = float(_editor_reference_cache.get("ts") or 0.0)
    if cached and now - cached_ts < _EDITOR_REFERENCE_CACHE_TTL_SECONDS:
        return cached

    dict_db = load_dictionaries_db()
    dict_items = []
    global_attrs = []
    for row in dict_db.get("items", []) or []:
        if not isinstance(row, dict):
            continue
        dict_id = str(row.get("id") or "").strip()
        if not dict_id:
            continue
        title = str(row.get("title") or dict_id).strip() or dict_id
        dict_items.append({"id": dict_id, "title": title, "size": len(row.get("items") or [])})
        global_attrs.append({
            "id": str(row.get("attr_id") or dict_id),
            "title": title,
            "code": str(row.get("code") or "").strip(),
            "type": _norm_type(row.get("type")),
            "scope": str(row.get("scope") or "both").strip() or "both",
            "dict_id": dict_id,
        })

    dict_items.sort(key=lambda x: str(x.get("title") or "").lower())
    global_attrs.sort(key=lambda x: str(x.get("title") or "").lower())
    payload = {"dict_items": dict_items, "attributes": global_attrs}
    _editor_reference_cache["ts"] = now
    _editor_reference_cache["payload"] = payload
    return payload


@router.get("/editor-reference")
def template_editor_reference() -> Dict[str, Any]:
    payload = _template_editor_reference_payload()
    return {"ok": True, **payload}


@router.get("/editor-bootstrap/{category_id}")
def template_editor_bootstrap(category_id: str) -> Dict[str, Any]:
    nodes = _get_catalog_nodes()
    path = _catalog_path(nodes, category_id)
    if not path:
        raise HTTPException(status_code=404, detail="CATEGORY_NOT_FOUND")

    path_ids = [str(item.get("id") or "").strip() for item in path if str(item.get("id") or "").strip()]
    payload = load_template_editor_payload(path_ids)
    owner_tpl = payload.get("template") if isinstance(payload.get("template"), dict) else None
    owner_attrs = _ensure_default_attrs(payload.get("attributes") if isinstance(payload.get("attributes"), list) else [])
    owner_category_id = str(payload.get("owner_category_id") or "").strip()
    owner_path_item = next((item for item in path if str(item.get("id") or "").strip() == owner_category_id), None)

    if isinstance(owner_tpl, dict):
        master = _template_master_payload(owner_tpl, owner_attrs, include_sources=False)
    else:
        master = {
            "version": 2,
            "base_attributes": _build_default_attrs(),
            "category_attributes": [],
            "stats": {
                "base_count": len(_build_default_attrs()),
                "category_count": 0,
                "required_count": sum(1 for x in _build_default_attrs() if bool(x.get("required"))),
                "total_count": len(_build_default_attrs()),
            },
            "sources": {},
        }

    own_template = owner_tpl if owner_path_item and str(owner_path_item.get("id") or "") == str(category_id) else None
    inherited_from = None if own_template else owner_path_item

    return {
        "ok": True,
        "category": {
            "id": str(category_id),
            "name": str(path[-1].get("name") or category_id),
            "path": path,
        },
        "owner_template": owner_tpl,
        "own_template": own_template,
        "inherited_from": inherited_from,
        "attributes": owner_attrs,
        "master": master,
    }


@router.post("/by-category/{category_id}")
def create_for_category(category_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    _editor_reference_cache["ts"] = 0.0
    _editor_reference_cache["payload"] = None
    """
    Создаёт НОВЫЙ шаблон для категории (теперь допускается множество).

    payload:
      { name?: string }
    """
    db = load_templates_db()

    tpl_id = new_id()
    name = (payload.get("name") or "Мастер-шаблон").strip()

    tpl = {
        "id": tpl_id,
        "category_id": category_id,
        "name": name,
        "updated_at": now_iso(),
        "created_at": now_iso(),
        "meta": {
            "sources": {},
        },
    }

    db.setdefault("templates", {})[tpl_id] = tpl
    db.setdefault("attributes", {})[tpl_id] = _build_default_attrs()

    # legacy mapping (не ломаем старые места)
    db.setdefault("category_to_template", {})
    if not db["category_to_template"].get(category_id):
        db["category_to_template"][category_id] = tpl_id

    # новый список
    db.setdefault("category_to_templates", {})
    db["category_to_templates"].setdefault(category_id, [])
    if tpl_id not in db["category_to_templates"][category_id]:
        db["category_to_templates"][category_id].append(tpl_id)

    save_templates_db(db)
    return {"template": tpl, "attributes": db.get("attributes", {}).get(tpl_id, []) or []}


@router.post("/by-category/{category_id}/apply-to-products")
def apply_template_to_products(category_id: str, payload: ApplyTemplateToProductsReq) -> Dict[str, Any]:
    db = load_templates_db()
    templates = db.get("templates") if isinstance(db.get("templates"), dict) else {}
    attrs_by_tpl = db.get("attributes") if isinstance(db.get("attributes"), dict) else {}
    cat_to_tpls = db.get("category_to_templates") if isinstance(db.get("category_to_templates"), dict) else {}

    tids = cat_to_tpls.get(category_id) if isinstance(cat_to_tpls.get(category_id), list) else []
    template_id = str((tids or [""])[0] or "").strip()
    tpl = templates.get(template_id) if template_id else None
    if not isinstance(tpl, dict):
        raise HTTPException(status_code=404, detail="TEMPLATE_NOT_FOUND")

    attrs = attrs_by_tpl.get(template_id) if isinstance(attrs_by_tpl.get(template_id), list) else []
    skeleton = _feature_skeleton_attrs(attrs)
    nodes = _load_catalog_nodes()
    target_category_ids = _descendant_category_ids(nodes, category_id) if payload.include_descendants else [str(category_id or "").strip()]
    target_set = {cid for cid in target_category_ids if cid}

    lock = with_lock("products_write")
    lock.acquire()
    try:
        products_doc = _load_products_doc()
        items = products_doc.get("items") if isinstance(products_doc.get("items"), list) else []
        matched = 0
        updated = 0
        changed_items: List[Dict[str, Any]] = []
        for product in items:
            if not isinstance(product, dict):
                continue
            cid = str(product.get("category_id") or "").strip()
            if cid not in target_set:
                continue
            matched += 1
            content = product.get("content") if isinstance(product.get("content"), dict) else {}
            merged_features = _merge_feature_skeleton(content.get("features"), skeleton)
            if merged_features != content.get("features"):
                updated += 1
                if not payload.dry_run:
                    next_content = dict(content)
                    next_content["features"] = merged_features
                    product["content"] = next_content
                    product["updated_at"] = now_iso()
                    changed_items.append(product)
        if not payload.dry_run:
            bulk_upsert_product_items(changed_items)
    finally:
        lock.release()

    return {
        "ok": True,
        "template_id": template_id,
        "category_id": category_id,
        "matched_products": matched,
        "updated_products": updated,
        "skeleton_fields": len(skeleton),
        "dry_run": bool(payload.dry_run),
        "include_descendants": bool(payload.include_descendants),
    }


@router.delete("/by-category/{category_id}")
def delete_for_category(category_id: str) -> Dict[str, Any]:
    """
    ⚠️ Старый endpoint удалял шаблон категории (когда он был один).
    Теперь безопаснее: удаляем legacy mapping, но НЕ удаляем все шаблоны,
    потому что их может быть много.
    """
    db = load_templates_db()
    db.get("category_to_template", {}).pop(category_id, None)
    db.get("category_to_templates", {}).pop(category_id, None)
    save_templates_db(db)
    return {"ok": True}


# =========================
# XLSX Export/Import
# =========================
XLSX_HEADERS_RU = ["Название", "Код", "Тип", "Скоуп", "Обязательный", "Опции"]
XLSX_HEADERS_EN = ["name", "code", "type", "scope", "required", "options"]

_HEADER_ALIASES = {
    "name": "name",
    "название": "name",
    "имя": "name",
    "code": "code",
    "код": "code",
    "ключ": "code",
    "type": "type",
    "тип": "type",
    "scope": "scope",
    "скоуп": "scope",
    "область": "scope",
    "уровень": "scope",
    "required": "required",
    "обязательный": "required",
    "обяз.": "required",
    "обяз": "required",
    "options": "options",
    "опции": "options",
    "варианты": "options",
    "значения": "options",
}


def _norm_header_cell(v: Any) -> str:
    s = str(v or "").strip().lower()
    s = s.replace("ё", "е")
    return s


def _build_header_map(vals: List[str]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for idx, raw in enumerate(vals, start=1):
        key = _HEADER_ALIASES.get(raw)
        if key and key not in col_map:
            col_map[key] = idx
    return col_map


def _has_all_required_cols(col_map: Dict[str, int]) -> bool:
    need = {"name", "code", "type", "scope", "required", "options"}
    return need.issubset(set(col_map.keys()))


def _xlsx_row_from_attr(a: Dict[str, Any]) -> List[Any]:
    a_type = _norm_type(a.get("type"))
    a_scope = _norm_scope(a.get("scope"))
    required = bool(a.get("required"))

    options = a.get("options") or {}
    if a_type == "select":
        vals = options.get("values") or []
        if isinstance(vals, list):
            options_str = ", ".join([str(x) for x in vals if str(x).strip()])
        else:
            options_str = str(vals)
    else:
        options_str = ""

    return [
        a.get("name") or "",
        a.get("code") or "",
        TYPE_RU.get(a_type, a_type),
        SCOPE_RU.get(a_scope, a_scope),
        "да" if required else "",
        options_str,
    ]


def _attr_from_xlsx_row(row: Dict[str, Any], position: int) -> Dict[str, Any]:
    name = (row.get("name") or "").strip()
    if not name:
        return {}

    code = (row.get("code") or "").strip().lower()
    if not code:
        code = slugify_code(name)

    a_type = _norm_type(row.get("type"))
    a_scope = _norm_scope(row.get("scope"))
    required = _boolish(row.get("required"))

    options_raw = row.get("options")
    options: Dict[str, Any] = {}
    if a_type == "select":
        s = (str(options_raw or "")).strip()
        values = [x.strip() for x in s.split(",") if x.strip()]
        options = {"values": values}

    return {
        "id": new_id(),
        "name": name,
        "code": code,
        "type": a_type,
        "required": required,
        "scope": a_scope,
        "options": options,
        "position": position,
        "locked": False,
    }


@router.get("/by-category/{category_id}/export.xlsx")
def export_template_xlsx(category_id: str):
    """
    ⚠️ legacy: экспорт для категории. Если шаблонов много — экспортируем первый (самый свежий).
    """
    from openpyxl import Workbook

    db = load_templates_db()
    cat_map = _templates_by_category(db)
    tids = cat_map.get(category_id, []) or []
    tpl_id = tids[0] if tids else None

    tpl = db.get("templates", {}).get(tpl_id) if tpl_id else None
    attrs = (db.get("attributes", {}).get(tpl_id, []) if tpl_id else []) or []
    attrs = _ensure_default_attrs(attrs)

    wb = Workbook()
    ws = wb.active
    ws.title = "template"

    ws["A1"] = "template_name"
    ws["B1"] = (tpl.get("name") if tpl else "").strip()

    ws.append([])
    ws.append(XLSX_HEADERS_RU)

    for a in sorted(attrs, key=lambda x: int(x.get("position", 0))):
        ws.append(_xlsx_row_from_attr(a))

    for col in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 22

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"template_{category_id}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/by-category/{category_id}/import.xlsx")
async def import_template_xlsx(category_id: str, file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Парсит Excel и возвращает attributes для фронта.
    ВАЖНО: здесь НЕ сохраняем — сохранение делается кнопкой "Сохранить".
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload .xlsx file")

    from openpyxl import load_workbook

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    template_name = ""
    try:
        if str(ws["A1"].value or "").strip().lower() == "template_name":
            template_name = str(ws["B1"].value or "").strip()
    except Exception:
        template_name = ""

    header_row_idx = None
    header_map: Dict[str, int] = {}

    for r in range(1, min(ws.max_row, 60) + 1):
        vals = [_norm_header_cell(ws.cell(row=r, column=c).value) for c in range(1, 30)]
        col_map = _build_header_map(vals)
        if _has_all_required_cols(col_map):
            header_row_idx = r
            header_map = col_map
            break

    if not header_row_idx:
        raise HTTPException(
            status_code=400,
            detail=f"Header row not found. Expected RU headers: {XLSX_HEADERS_RU} (or EN: {XLSX_HEADERS_EN})",
        )

    parsed: List[Dict[str, Any]] = []
    pos = 0

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_obj: Dict[str, Any] = {}
        empty = True

        for h in ["name", "code", "type", "scope", "required", "options"]:
            c = header_map[h]
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                empty = False
            row_obj[h] = "" if v is None else str(v).strip()

        if empty:
            continue

        a = _attr_from_xlsx_row(row_obj, pos)
        if a:
            parsed.append(a)
            pos += 1

    parsed = _dedupe_codes(parsed)
    parsed = _ensure_default_attrs(parsed)

    return {
        "ok": True,
        "template_name": template_name,
        "attributes": parsed,
    }


# ============================================================
# ✅ PARAM ROUTES LAST (ВАЖНО: ПОСЛЕ /tree и т.п.)
# ============================================================

@router.get("/{template_id}")
def get_template(template_id: str) -> Dict[str, Any]:
    """
    ✅ Удобный эндпоинт: получить шаблон + атрибуты по template_id.
    """
    db = load_templates_db()
    tpl = (db.get("templates") or {}).get(template_id)
    if not tpl:
        raise HTTPException(status_code=404, detail="template not found")
    attrs = (db.get("attributes") or {}).get(template_id, []) or []
    attrs = _ensure_default_attrs(attrs)
    master = _template_master_payload(tpl, attrs)
    return {
        "ok": True,
        "template": tpl,
        "attributes": attrs,
        "base_attributes": master["base_attributes"],
        "category_attributes": master["category_attributes"],
        "master": master,
    }


@router.put("/{template_id}")
def update_template(template_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    _editor_reference_cache["ts"] = 0.0
    _editor_reference_cache["payload"] = None
    db = load_templates_db()
    tpl = db.get("templates", {}).get(template_id)
    if not tpl:
        raise HTTPException(status_code=404, detail="template not found")

    if "name" in payload:
        nm = (payload.get("name") or "").strip()
        if nm:
            tpl["name"] = nm

    tpl["updated_at"] = now_iso()
    db["templates"][template_id] = tpl
    save_templates_db(db)
    return {"template": tpl}


@router.put("/{template_id}/attributes")
def replace_attributes(template_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    _editor_reference_cache["ts"] = 0.0
    _editor_reference_cache["payload"] = None
    """
    payload: { attributes: [...] }
    полностью перезаписываем список атрибутов

    ✅ ВАЖНО:
    - Для параметров типа "select" автоматически создаём словарь в backend/data/dictionaries.json
    - И сохраняем dict_id внутрь options (options.dict_id)
    """
    db = load_templates_db()
    tpl = db.get("templates", {}).get(template_id)
    if not tpl:
        raise HTTPException(status_code=404, detail="template not found")

    attrs_in = payload.get("attributes") or []
    out = _normalize_attributes(attrs_in)
    out = _ensure_default_attrs(out)

    try:
        from app.api.routes.dictionaries import ensure_dictionaries_for_template_attrs

        ensure_dictionaries_for_template_attrs(out)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"dictionaries sync failed: {e}")

    db.setdefault("attributes", {})[template_id] = out
    tpl["updated_at"] = now_iso()
    tpl_meta = tpl.get("meta") if isinstance(tpl.get("meta"), dict) else {}
    master_payload = _template_master_payload(tpl, out)
    tpl_meta["master_template"] = {
        "version": master_payload["version"],
        "stats": master_payload["stats"],
    }
    tpl["meta"] = tpl_meta
    db["templates"][template_id] = tpl
    save_templates_db(db)

    return {"ok": True, "attributes": out}


@router.delete("/{template_id}")
def delete_template(template_id: str) -> Dict[str, Any]:
    _editor_reference_cache["ts"] = 0.0
    _editor_reference_cache["payload"] = None
    """
    ✅ Удаление конкретного шаблона по template_id.
    """
    db = load_templates_db()
    tpl = db.get("templates", {}).get(template_id)
    if not tpl:
        return {"ok": True}

    category_id = str(tpl.get("category_id") or "")

    db.get("templates", {}).pop(template_id, None)
    db.get("attributes", {}).pop(template_id, None)

    # подчистим новые/старые mapping-и
    if category_id:
        cat_to_tpls = db.get("category_to_templates", {}) or {}
        if isinstance(cat_to_tpls, dict) and isinstance(cat_to_tpls.get(category_id), list):
            cat_to_tpls[category_id] = [x for x in cat_to_tpls[category_id] if str(x) != str(template_id)]
            db["category_to_templates"] = cat_to_tpls

        cat_to_tpl = db.get("category_to_template", {}) or {}
        if isinstance(cat_to_tpl, dict) and str(cat_to_tpl.get(category_id) or "") == str(template_id):
            left = (db.get("category_to_templates", {}) or {}).get(category_id) or []
            cat_to_tpl[category_id] = left[0] if left else None
            if not cat_to_tpl[category_id]:
                cat_to_tpl.pop(category_id, None)
            db["category_to_template"] = cat_to_tpl

    save_templates_db(db)
    return {"ok": True}
