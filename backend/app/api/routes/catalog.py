from __future__ import annotations

import uuid
import time
from pathlib import Path
from typing import Optional, List, Dict, Any, Set, Tuple

from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from pydantic import BaseModel, Field
from fastapi.responses import StreamingResponse

from app.storage.json_store import load_templates_db, save_templates_db
from app.core.json_store import read_doc, write_doc
from app.core.products.service import (
    create_product_service,
    allocate_sku_pairs_service,
    delete_products_bulk_service,
)
from app.api.routes.products import (
    YANDEX_OFFER_CARDS_PATH,
    OZON_PRODUCT_RATING_PATH,
    OZON_IMPORT_INFO_PATH,
    _normalize_market_status,
    _normalize_ozon_status,
    _offer_ids_for_product,
    _load_connectors_state,
)

router = APIRouter(tags=["Catalog"])

BASE_DIR = Path(__file__).resolve().parents[3]  # backend/
DATA_DIR = BASE_DIR / "data"

CATALOG_PATH = DATA_DIR / "catalog_nodes.json"
PRODUCTS_PATH = DATA_DIR / "catalog_products.json"
FULL_PRODUCTS_PATH = DATA_DIR / "products.json"

_PRODUCTS_PAGE_CACHE_TTL_SECONDS = 20.0
_products_page_cache: Dict[str, Any] = {"ts": 0.0, "payload": None}


# =========================
# IO helpers
# =========================
def _load_json(path: Path, default):
    return read_doc(path, default=default)


def _save_json(path: Path, data) -> None:
    write_doc(path, data)


def _load_nodes() -> List[Dict[str, Any]]:
    return _load_json(CATALOG_PATH, [])


def _save_nodes(nodes: List[Dict[str, Any]]) -> None:
    _save_json(CATALOG_PATH, nodes)


def _load_products() -> List[Dict[str, Any]]:
    return _load_json(PRODUCTS_PATH, [])


def _save_products(items: List[Dict[str, Any]]) -> None:
    _save_json(PRODUCTS_PATH, items)


def _load_full_products() -> List[Dict[str, Any]]:
    doc = _load_json(FULL_PRODUCTS_PATH, {"items": []})
    items = doc.get("items") if isinstance(doc, dict) else None
    return items if isinstance(items, list) else []


def _save_full_products(items: List[Dict[str, Any]]) -> None:
    _save_json(FULL_PRODUCTS_PATH, {"items": items})


def _serialize_product_list_item(product: Dict[str, Any]) -> Dict[str, Any]:
    title = str(product.get("title") or product.get("name") or "")
    content = product.get("content") if isinstance(product.get("content"), dict) else {}
    media_images = content.get("media_images") if isinstance(content.get("media_images"), list) else []
    media_legacy = content.get("media") if isinstance(content.get("media"), list) else []
    media_pool = media_images if media_images else media_legacy
    preview_url = ""
    for item in media_pool:
        if isinstance(item, dict) and str(item.get("url") or "").strip():
            preview_url = str(item.get("url") or "").strip()
            break
    return {
        "id": str(product.get("id") or ""),
        "name": title,
        "title": title,
        "category_id": str(product.get("category_id") or ""),
        "sku_pim": str(product.get("sku_pim") or ""),
        "sku_gt": str(product.get("sku_gt") or ""),
        "group_id": str(product.get("group_id") or ""),
        "preview_url": preview_url,
        "exports_enabled": product.get("exports_enabled") if isinstance(product.get("exports_enabled"), dict) else {},
    }


def _products_page_meta() -> Dict[str, Any]:
    now = time.monotonic()
    cached_payload = _products_page_cache.get("payload")
    cached_ts = float(_products_page_cache.get("ts") or 0.0)
    if cached_payload and now - cached_ts < _PRODUCTS_PAGE_CACHE_TTL_SECONDS:
        return cached_payload

    nodes = _load_nodes()
    groups_doc = read_doc(DATA_DIR / "product_groups.json", default={"items": []})
    group_items = groups_doc.get("items") if isinstance(groups_doc, dict) else []
    groups = group_items if isinstance(group_items, list) else []

    templates_db = load_templates_db()
    templates_map = templates_db.get("templates") if isinstance(templates_db.get("templates"), dict) else {}
    template_items = []
    for tid, row in templates_map.items():
        if not isinstance(row, dict):
            continue
        template_items.append(
            {
                "id": str(row.get("id") or tid),
                "category_id": str(row.get("category_id") or ""),
                "name": str(row.get("name") or ""),
            }
        )
    template_items.sort(key=lambda x: x["name"].lower())

    payload = {
        "nodes": nodes,
        "groups": groups,
        "templates": template_items,
        "templates_db": templates_db,
    }
    _products_page_cache["ts"] = now
    _products_page_cache["payload"] = payload
    return payload


def _gt_sort_key(value: Any) -> Tuple[int, int, str]:
    v = str(value or "").strip()
    if not v:
        return 1, 2**31 - 1, ""
    if v.isdigit():
        return 0, int(v), v
    return 0, 2**31 - 1, v.lower()


def _build_marketplace_status_context() -> Dict[str, Any]:
    state = _load_connectors_state()
    providers = state.get("providers") if isinstance(state, dict) else {}
    if not isinstance(providers, dict):
        providers = {}

    yandex_provider = providers.get("yandex_market") if isinstance(providers.get("yandex_market"), dict) else {}
    ozon_provider = providers.get("ozon") if isinstance(providers.get("ozon"), dict) else {}

    yandex_stores = yandex_provider.get("import_stores") if isinstance(yandex_provider.get("import_stores"), list) else []
    ozon_stores = ozon_provider.get("import_stores") if isinstance(ozon_provider.get("import_stores"), list) else []

    cards_doc = read_doc(YANDEX_OFFER_CARDS_PATH, default={"items": {}})
    card_items = cards_doc.get("items") if isinstance(cards_doc, dict) else {}
    if not isinstance(card_items, dict):
        card_items = {}

    rating_doc = read_doc(OZON_PRODUCT_RATING_PATH, default={"items": {}})
    rating_items = rating_doc.get("items") if isinstance(rating_doc, dict) else {}
    if not isinstance(rating_items, dict):
        rating_items = {}

    import_doc = read_doc(OZON_IMPORT_INFO_PATH, default={"items": {}})
    import_items = import_doc.get("items") if isinstance(import_doc, dict) else {}
    if not isinstance(import_items, dict):
        import_items = {}

    return {
        "yandex_stores": yandex_stores,
        "ozon_stores": ozon_stores,
        "card_items": card_items,
        "rating_items": rating_items,
        "import_items": import_items,
    }


def _marketplace_statuses_for_product(product: Dict[str, Any], ctx: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    offer_ids = _offer_ids_for_product(product)

    yandex_stores_by_id: Dict[str, Dict[str, Any]] = {}
    for store in ctx.get("yandex_stores") or []:
        if not isinstance(store, dict):
            continue
        store_id = str(store.get("id") or "").strip()
        if not store_id:
            continue
        yandex_stores_by_id[store_id] = {
            "store_id": store_id,
            "status_code": "",
            "status": "Нет данных",
        }

    for offer_id in offer_ids:
        row = ctx.get("card_items", {}).get(offer_id)
        if not isinstance(row, dict):
            continue
        sources = row.get("sources") if isinstance(row.get("sources"), dict) else {}
        for source_key, src in sources.items():
            if not isinstance(src, dict):
                continue
            store_id = str(src.get("store_id") or source_key or "").strip()
            if not store_id or store_id.isdigit():
                continue
            card = src.get("card") if isinstance(src.get("card"), dict) else {}
            status_code = str(card.get("cardStatus") or "").strip()
            yandex_stores_by_id[store_id] = {
                "store_id": store_id,
                "status_code": status_code,
                "status": _normalize_market_status(status_code),
            }

    yandex_stores = list(yandex_stores_by_id.values())
    yandex_codes = {str(x.get("status_code") or "") for x in yandex_stores if str(x.get("status_code") or "").strip()}
    yandex_summary = "Нет данных" if not yandex_stores else ("Есть расхождения" if len(yandex_codes) > 1 else (yandex_stores[0].get("status") or "Нет данных"))
    yandex_present = any(str(x.get("status_code") or "").strip() for x in yandex_stores)

    ozon_rows: List[Dict[str, Any]] = []
    for store in ctx.get("ozon_stores") or []:
        if not isinstance(store, dict):
            continue
        store_id = str(store.get("id") or "").strip()
        if not store_id:
            continue
        import_row = None
        rating_row = None
        for offer_id in offer_ids:
            maybe_import = ctx.get("import_items", {}).get(f"{store_id}:{offer_id}")
            if isinstance(maybe_import, dict):
                import_row = maybe_import
            maybe_rating = ctx.get("rating_items", {}).get(f"{store_id}:{offer_id}")
            if isinstance(maybe_rating, dict):
                rating_row = maybe_rating
            if import_row or rating_row:
                break
        status_code = str((import_row or {}).get("status") or (import_row or {}).get("state") or "").strip()
        ozon_rows.append(
            {
                "store_id": store_id,
                "status_code": status_code,
                "status": (str((import_row or {}).get("status") or "").strip() or _normalize_ozon_status(status_code) or "Нет данных"),
                "has_rating": bool(rating_row),
            }
        )

    ozon_codes = {str(x.get("status_code") or "") for x in ozon_rows if str(x.get("status_code") or "").strip()}
    ozon_summary = "Нет данных" if not ozon_rows else ("Есть расхождения" if len(ozon_codes) > 1 else (ozon_rows[0].get("status") or "Нет данных"))
    ozon_present = any(str(x.get("status_code") or "").strip() or bool(x.get("has_rating")) for x in ozon_rows)

    return {
        "yandex_market": {"status": yandex_summary, "present": yandex_present},
        "ozon": {"status": ozon_summary, "present": ozon_present},
    }


def _cleanup_templates_for_deleted_categories(deleted_category_ids: set[str]) -> None:
    """
    Удаляет шаблоны/маппинги, привязанные к удалённым категориям.
    """
    if not deleted_category_ids:
        return

    db = load_templates_db()
    templates = db.get("templates") if isinstance(db.get("templates"), dict) else {}
    attrs = db.get("attributes") if isinstance(db.get("attributes"), dict) else {}
    cat_to_tpl = db.get("category_to_template") if isinstance(db.get("category_to_template"), dict) else {}
    cat_to_tpls = db.get("category_to_templates") if isinstance(db.get("category_to_templates"), dict) else {}

    removed_tids: set[str] = set()
    changed = False

    for tid, t in list(templates.items()):
        if not isinstance(t, dict):
            templates.pop(tid, None)
            removed_tids.add(str(tid))
            changed = True
            continue
        cid = str(t.get("category_id") or "").strip()
        if cid and cid in deleted_category_ids:
            templates.pop(tid, None)
            removed_tids.add(str(tid))
            changed = True

    for tid in removed_tids:
        if tid in attrs:
            attrs.pop(tid, None)
            changed = True

    for cid in list(cat_to_tpl.keys()):
        tid = str(cat_to_tpl.get(cid) or "").strip()
        if cid in deleted_category_ids or (tid and tid in removed_tids):
            cat_to_tpl.pop(cid, None)
            changed = True

    for cid in list(cat_to_tpls.keys()):
        if cid in deleted_category_ids:
            cat_to_tpls.pop(cid, None)
            changed = True
            continue
        arr = cat_to_tpls.get(cid)
        if not isinstance(arr, list):
            cat_to_tpls.pop(cid, None)
            changed = True
            continue
        next_arr = [str(x) for x in arr if str(x) and str(x) not in removed_tids and str(x) in templates]
        if next_arr != arr:
            if next_arr:
                cat_to_tpls[cid] = next_arr
            else:
                cat_to_tpls.pop(cid, None)
            changed = True

    if changed:
        db["templates"] = templates
        db["attributes"] = attrs
        db["category_to_template"] = cat_to_tpl
        db["category_to_templates"] = cat_to_tpls
        save_templates_db(db)


# =========================
# Tree helpers
# =========================
def _normalize_positions(nodes: List[Dict[str, Any]], parent_id: Optional[str]) -> None:
    siblings = [n for n in nodes if (n.get("parent_id") or None) == (parent_id or None)]
    siblings.sort(key=lambda x: int(x.get("position", 0)))
    for i, n in enumerate(siblings):
        n["position"] = i


def _is_descendant(nodes: List[Dict[str, Any]], node_id: str, maybe_parent_id: str) -> bool:
    # Проверяем: maybe_parent_id находится в поддереве node_id?
    by_id = {n["id"]: n for n in nodes}
    cur = by_id.get(maybe_parent_id)
    while cur:
        pid = cur.get("parent_id")
        if pid == node_id:
            return True
        cur = by_id.get(pid) if pid else None
    return False


def _collect_subtree_ids(nodes: List[Dict[str, Any]], root_id: str) -> set[str]:
    by_parent: Dict[Optional[str], List[str]] = {}
    for n in nodes:
        by_parent.setdefault(n.get("parent_id") or None, []).append(n["id"])

    to_delete: set[str] = set()

    def dfs(x: str):
        to_delete.add(x)
        for child_id in by_parent.get(x, []):
            if child_id not in to_delete:
                dfs(child_id)

    dfs(root_id)
    return to_delete


def _attach_products_count(nodes: List[Dict[str, Any]], products: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    products_count — прямое количество товаров в категории (не агрегированное).
    Фронт сам агрегирует (computeAggregatedCounts).
    """
    counts: Dict[str, int] = {}
    for p in products:
        cid = p.get("category_id")
        if cid:
            counts[cid] = counts.get(cid, 0) + 1

    out: List[Dict[str, Any]] = []
    for n in nodes:
        nn = dict(n)
        nn["products_count"] = int(counts.get(nn["id"], 0))
        # на всякий случай поля по умолчанию
        nn.setdefault("template_id", None)
        nn.setdefault("position", 0)
        nn.setdefault("parent_id", None)
        out.append(nn)

    return out


def _build_category_path(nodes: List[Dict[str, Any]], category_id: str) -> List[Dict[str, str]]:
    by_id: Dict[str, Dict[str, Any]] = {str(n.get("id")): n for n in nodes}
    cur = by_id.get(str(category_id))
    if not cur:
        return []
    chain: List[Dict[str, str]] = []
    seen = set()
    while cur:
        cid = str(cur.get("id"))
        if cid in seen:
            break
        seen.add(cid)
        chain.append({"id": cid, "name": str(cur.get("name") or "")})
        pid = cur.get("parent_id") or None
        cur = by_id.get(str(pid)) if pid else None
    chain.reverse()
    return chain


def _templates_by_category(db: Dict[str, Any]) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    templates = db.get("templates") or {}

    if isinstance(templates, dict):
        for tid, t in templates.items():
            if not isinstance(t, dict):
                continue
            cid = str(t.get("category_id") or "").strip()
            if not cid:
                continue
            out.setdefault(cid, [])
            if tid not in out[cid]:
                out[cid].append(tid)

    cat_to_tpls = db.get("category_to_templates") or {}
    if isinstance(cat_to_tpls, dict):
        for cid, arr in cat_to_tpls.items():
            if not isinstance(arr, list):
                continue
            out.setdefault(cid, [])
            for tid in arr:
                if tid not in out[cid]:
                    out[cid].append(tid)

    return out


def _resolve_template_for_category(
    nodes: List[Dict[str, Any]],
    category_id: str,
    cat_map: Dict[str, List[str]],
) -> Tuple[Optional[str], Optional[str]]:
    by_id = {n.get("id"): n for n in nodes}
    cur = by_id.get(category_id)
    while cur:
        cid = cur.get("id")
        tids = cat_map.get(cid, []) or []
        if tids:
            return tids[0], cid
        pid = cur.get("parent_id")
        cur = by_id.get(pid) if pid else None
    return None, None


# =========================
# DTOs
# =========================
class NodeCreate(BaseModel):
    name: str = Field(min_length=1)
    parent_id: Optional[str] = None


class NodePatch(BaseModel):
    name: Optional[str] = None
    position: Optional[int] = None


class NodeMove(BaseModel):
    new_parent_id: Optional[str] = None
    new_position: int = 0


class ProductCreate(BaseModel):
    name: str = Field(min_length=1)
    category_id: str = Field(min_length=1)


class ProductDeleteReq(BaseModel):
    ids: List[str] = Field(default_factory=list)


# =========================
# Nodes endpoints
# =========================
@router.get("/catalog/nodes")
def list_nodes():
    return get_nodes_json()

@router.post("/catalog/nodes")
def create_node(payload: NodeCreate):
    nodes = _load_nodes()

    node_id = str(uuid.uuid4())
    parent_id = payload.parent_id

    # позиция = в конец siblings
    siblings = [n for n in nodes if (n.get("parent_id") or None) == (parent_id or None)]
    position = len(siblings)

    node = {
        "id": node_id,
        "parent_id": parent_id,
        "name": payload.name.strip(),
        "position": position,
        "template_id": None,
        "products_count": 0,  # будет пересчитано в list_nodes
    }
    nodes.append(node)
    _save_nodes(nodes)
    return node


@router.patch("/catalog/nodes/{node_id}")
def patch_node(node_id: str, payload: NodePatch):
    nodes = _load_nodes()
    target: Optional[Dict[str, Any]] = None
    for n in nodes:
        if n["id"] == node_id:
            target = n
            if payload.name is not None:
                n["name"] = payload.name.strip()
            break
    if not target:
        raise HTTPException(status_code=404, detail="Node not found")

    if payload.position is not None:
        parent_id = target.get("parent_id") or None
        siblings = [n for n in nodes if (n.get("parent_id") or None) == parent_id and n["id"] != node_id]
        siblings.sort(key=lambda x: int(x.get("position", 0)))
        pos = max(0, min(int(payload.position), len(siblings)))
        merged = siblings[:pos] + [target] + siblings[pos:]
        for i, node in enumerate(merged):
            node["position"] = i

    _save_nodes(nodes)
    return target


@router.delete("/catalog/nodes/{node_id}")
def delete_node(node_id: str):
    nodes = _load_nodes()
    by_id = {n["id"]: n for n in nodes}
    if node_id not in by_id:
        raise HTTPException(status_code=404, detail="Node not found")

    # 1) собираем subtree
    subtree_ids = _collect_subtree_ids(nodes, node_id)

    # 2) удаляем узлы
    new_nodes = [n for n in nodes if n["id"] not in subtree_ids]

    # 3) нормализуем позиции у всех родителей
    parents = set((n.get("parent_id") or None) for n in new_nodes)
    for pid in parents:
        _normalize_positions(new_nodes, pid)

    _save_nodes(new_nodes)

    # 4) удаляем товары, которые лежат в удаленной ветке
    products = _load_products()
    new_products = [p for p in products if p.get("category_id") not in subtree_ids]
    if len(new_products) != len(products):
        _save_products(new_products)

    # 5) удаляем шаблоны/привязки для удалённых категорий
    _cleanup_templates_for_deleted_categories(subtree_ids)

    return {"ok": True}


@router.patch("/catalog/nodes/{node_id}/move")
def move_node(node_id: str, payload: NodeMove):
    nodes = _load_nodes()
    by_id = {n["id"]: n for n in nodes}
    node = by_id.get(node_id)
    if not node:
        raise HTTPException(status_code=404, detail="Node not found")

    new_parent_id = payload.new_parent_id
    if new_parent_id == node_id:
        raise HTTPException(status_code=400, detail="Cannot move into itself")

    if new_parent_id and _is_descendant(nodes, node_id, new_parent_id):
        raise HTTPException(status_code=400, detail="Cannot move into descendant")

    old_parent_id = node.get("parent_id") or None

    # удаляем из старых siblings
    old_siblings = [
        n for n in nodes
        if (n.get("parent_id") or None) == (old_parent_id or None) and n["id"] != node_id
    ]
    old_siblings.sort(key=lambda x: int(x.get("position", 0)))

    # новые siblings
    new_siblings = [
        n for n in nodes
        if (n.get("parent_id") or None) == (new_parent_id or None) and n["id"] != node_id
    ]
    new_siblings.sort(key=lambda x: int(x.get("position", 0)))

    # вставка в new_position
    pos = max(0, min(payload.new_position, len(new_siblings)))
    node["parent_id"] = new_parent_id

    merged = new_siblings[:pos] + [node] + new_siblings[pos:]
    for i, n in enumerate(merged):
        n["position"] = i

    for i, n in enumerate(old_siblings):
        n["position"] = i

    _save_nodes(nodes)
    return {"ok": True}

def get_nodes_json():
    nodes = _load_nodes()
    products = _load_products()
    nodes = _attach_products_count(nodes, products)
    return {"nodes": nodes}

@router.get("/catalog/categories/{category_id}")
def get_category_info(category_id: str) -> Dict[str, Any]:
    """
    Для хлебных крошек.
    Возвращает: { id, name, path: [{id,name}, ...] }  (path от корня до текущей)
    """
    nodes = _load_nodes()
    by_id: Dict[str, Dict[str, Any]] = {str(n.get("id")): n for n in nodes}

    cur = by_id.get(str(category_id))
    if not cur:
        raise HTTPException(status_code=404, detail="Category not found")

    chain: List[Dict[str, str]] = []
    seen = set()

    # идём вверх по parent_id
    while cur:
        cid = str(cur.get("id"))
        if cid in seen:
            break
        seen.add(cid)

        chain.append({"id": cid, "name": str(cur.get("name") or "")})

        pid = cur.get("parent_id") or None
        cur = by_id.get(str(pid)) if pid else None

    chain.reverse()

    return {
        "id": str(category_id),
        "name": str(by_id[str(category_id)].get("name") or ""),
        "path": chain,
    }


# =========================
# Products endpoints (JSON-based)
# =========================
@router.get("/catalog/products/search")
def search_products(q: str = ""):
    query = (q or "").strip().lower()
    if not query:
        return {"items": []}

    products = _load_full_products()
    items = []
    for p in products:
        title = str(p.get("title") or p.get("name") or "")
        sku_pim = str(p.get("sku_pim") or "")
        sku_gt = str(p.get("sku_gt") or "")
        haystack = [
            title.lower(),
            sku_pim.lower(),
            sku_gt.lower(),
        ]
        if not any(query in x for x in haystack):
            continue

        items.append(
            {
                "id": str(p.get("id") or ""),
                "name": title,
                "category_id": str(p.get("category_id") or ""),
            }
        )

    # лимит чтобы не раздувать ответ
    return {"items": items[:50]}


@router.get("/catalog/products")
def list_products(category_id: Optional[str] = None, include_descendants: bool = True):
    products = _load_full_products()
    if category_id:
        if include_descendants:
            nodes = _load_nodes()
            subtree_ids = _collect_subtree_ids(nodes, category_id)
            products = [p for p in products if p.get("category_id") in subtree_ids]
        else:
            products = [p for p in products if p.get("category_id") == category_id]

    products.sort(
        key=lambda p: (
            _gt_sort_key(p.get("sku_gt")),
            str(p.get("title") or p.get("name") or "").lower(),
        )
    )

    return {"items": [_serialize_product_list_item(p) for p in products]}


@router.get("/catalog/products-page-data")
def products_page_data(
    q: str = Query(""),
    category: str = Query(""),
    exact: bool = Query(False),
    parent: str = Query(""),
    sub: str = Query(""),
    group: str = Query(""),
    template: str = Query(""),
    ym: str = Query("all"),
    oz: str = Query("all"),
    view: str = Query("all"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    refresh: bool = Query(False),
):
    if refresh:
        _products_page_cache["ts"] = 0.0
        _products_page_cache["payload"] = None

    meta = _products_page_meta()
    nodes = meta["nodes"]
    groups = meta["groups"]
    templates = meta["templates"]
    templates_db = meta["templates_db"]

    group_name_by_id = {str(g.get("id") or ""): str(g.get("name") or "") for g in groups if isinstance(g, dict)}
    templates_by_category = _templates_by_category(templates_db)
    exact_category = (category or "").strip()
    subtree_parent = _collect_subtree_ids(nodes, parent) if parent else None
    subtree_sub = _collect_subtree_ids(nodes, sub) if sub else None
    subtree_exact = _collect_subtree_ids(nodes, exact_category) if exact_category and not exact else None
    q_normalized = (q or "").strip().lower()
    ym_filter = (ym or "all").strip().lower()
    oz_filter = (oz or "all").strip().lower()
    view_filter = (view or "all").strip().lower()
    template_cache: Dict[str, Tuple[Optional[str], Optional[str]]] = {}
    path_cache: Dict[str, str] = {}

    def category_path(category_id: str) -> str:
        if category_id in path_cache:
            return path_cache[category_id]
        path = " / ".join(part["name"] for part in _build_category_path(nodes, category_id) if part.get("name"))
        path_cache[category_id] = path
        return path

    def resolved_template(category_id: str) -> Tuple[str, str, str]:
        if category_id in template_cache:
            tid, source_cid = template_cache[category_id]
        else:
            tid, source_cid = _resolve_template_for_category(nodes, category_id, templates_by_category)
            template_cache[category_id] = (tid, source_cid)
        if not tid:
            return "", "", ""
        template_name = ""
        for row in templates:
            if str(row.get("id") or "") == str(tid):
                template_name = str(row.get("name") or "")
                break
        return str(tid), template_name, str(source_cid or "")

    marketplace_ctx = _build_marketplace_status_context()
    filtered_items: List[Dict[str, Any]] = []
    for product in _load_full_products():
        item = _serialize_product_list_item(product)
        category_id = str(item.get("category_id") or "")
        if exact_category:
            if exact:
                if category_id != exact_category:
                    continue
            elif subtree_exact and category_id not in subtree_exact:
                continue
        if subtree_parent and category_id not in subtree_parent:
            continue
        if subtree_sub and category_id not in subtree_sub:
            continue

        group_id = str(item.get("group_id") or "").strip()
        if group == "__ungrouped__" and group_id:
            continue
        if group and group != "__ungrouped__" and group_id != group:
            continue

        template_id, template_name, template_source_category_id = resolved_template(category_id)
        if template == "__without__" and template_id:
            continue
        if template and template != "__without__" and template_id != template:
            continue

        marketplace_statuses = _marketplace_statuses_for_product(product, marketplace_ctx)
        export_ym = bool(((marketplace_statuses.get("yandex_market") or {}).get("present")))
        export_oz = bool(((marketplace_statuses.get("ozon") or {}).get("present")))
        if ym_filter == "on" and not export_ym:
            continue
        if ym_filter == "off" and export_ym:
            continue
        if oz_filter == "on" and not export_oz:
            continue
        if oz_filter == "off" and export_oz:
            continue

        if view_filter == "issues" and template_id and export_ym and export_oz:
            continue
        if view_filter == "no_template" and template_id:
            continue
        if view_filter == "no_ym" and export_ym:
            continue
        if view_filter == "no_oz" and export_oz:
            continue

        cat_path = category_path(category_id)
        if q_normalized:
            haystack = " ".join(
                [
                    str(item.get("title") or ""),
                    str(item.get("sku_gt") or ""),
                    cat_path,
                ]
            ).lower()
            if q_normalized not in haystack:
                continue

        item["category_path"] = cat_path
        item["group_name"] = group_name_by_id.get(group_id, "")
        item["effective_template_id"] = template_id
        item["effective_template_name"] = template_name
        item["effective_template_source_category_id"] = template_source_category_id
        item["marketplace_statuses"] = marketplace_statuses
        filtered_items.append(item)

    filtered_items.sort(
        key=lambda item: (
            _gt_sort_key(item.get("sku_gt"))[0],
            _gt_sort_key(item.get("sku_gt"))[1],
            _gt_sort_key(item.get("sku_gt"))[2],
            str(item.get("title") or item.get("name") or "").lower(),
        )
    )

    total = len(filtered_items)
    start = max(0, (page - 1) * page_size)
    end = start + page_size

    return {
        "ok": True,
        "products": filtered_items[start:end],
        "total": total,
        "page": page,
        "page_size": page_size,
        "nodes": nodes,
        "groups": groups,
        "templates": templates,
    }


@router.get("/catalog/products/template.xlsx")
def export_products_template(category_id: str = Query(..., min_length=1)):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    nodes = _load_nodes()
    by_id = {n.get("id"): n for n in nodes}
    if category_id not in by_id:
        raise HTTPException(status_code=404, detail="Category not found")

    db = load_templates_db()
    cat_map = _templates_by_category(db)
    tpl_id, tpl_category_id = _resolve_template_for_category(nodes, category_id, cat_map)
    tpl = (db.get("templates") or {}).get(tpl_id) if tpl_id else None
    attrs = (db.get("attributes") or {}).get(tpl_id, []) if tpl_id else []

    base_category_id = tpl_category_id or category_id

    children_map: Dict[str, List[Dict[str, Any]]] = {}
    for n in nodes:
        pid = n.get("parent_id") or None
        children_map.setdefault(pid, []).append(n)
    for k, arr in children_map.items():
        arr.sort(key=lambda x: str(x.get("name") or ""))
        children_map[k] = arr

    levels: Dict[int, List[str]] = {}
    stack: List[Tuple[str, int]] = [(base_category_id, 0)]
    seen = set()
    while stack:
        nid, depth = stack.pop(0)
        if nid in seen:
            continue
        seen.add(nid)
        node = by_id.get(nid)
        if node:
            name = str(node.get("name") or "").strip()
            if name:
                levels.setdefault(depth, [])
                if name not in levels[depth]:
                    levels[depth].append(name)
        for child in children_map.get(nid, []) or []:
            cid = child.get("id")
            if cid:
                stack.append((cid, depth + 1))

    max_depth = max(levels.keys()) if levels else 0
    for depth in levels:
        levels[depth].sort()

    wb = Workbook()
    ws = wb.active
    ws.title = "products"

    header_categories_ru = [f"Категория {i}" for i in range(1, max_depth + 2)]
    header_categories = [f"category_{i}" for i in range(1, max_depth + 2)]
    header_skus_ru = ["ID PIM", "ID GT"]
    header_skus = ["sku_pim", "sku_gt"]
    header_base_ru = ["Название товара"]
    header_base = ["title"]

    attr_headers_ru: List[str] = []
    attr_headers: List[str] = []
    attr_defs: List[Dict[str, Any]] = []
    seen_attrs: Set[str] = set()
    idx = 1
    for a in sorted(attrs or [], key=lambda x: int(x.get("position", 0))):
        name = str(a.get("name") or "").strip()
        code = str(a.get("code") or "").strip()
        if not name and not code:
            continue
        if not code:
            code = f"attr_{idx}"
        if code.lower() in seen_attrs:
            i = 2
            while f"{code} ({i})".lower() in seen_attrs:
                i += 1
            code = f"{code} ({i})"
        seen_attrs.add(code.lower())
        attr_headers_ru.append(name or code)
        attr_headers.append(code)
        attr_defs.append(a)
        idx += 1

    ws.append(header_categories_ru + header_skus_ru + header_base_ru + attr_headers_ru)
    ws.append(header_categories + header_skus + header_base + attr_headers)

    for idx in range(1, len(header_categories) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 28
    base_start = len(header_categories) + 1
    ws.column_dimensions[get_column_letter(base_start)].width = 16  # sku_pim
    ws.column_dimensions[get_column_letter(base_start + 1)].width = 16  # sku_gt
    ws.column_dimensions[get_column_letter(base_start + 2)].width = 40  # title
    for i in range(base_start + 3, base_start + 3 + len(attr_headers)):
        ws.column_dimensions[get_column_letter(i)].width = 24

    list_col = base_start + 3 + len(attr_headers) + 1
    data_start_row = 3
    data_end_row = 1000

    def _add_validation(target_col: int, values: List[str]) -> None:
        nonlocal list_col
        values = [v for v in values if v]
        if not values:
            return
        col_letter = get_column_letter(list_col)
        for r_idx, val in enumerate(values, start=1):
            ws.cell(row=r_idx, column=list_col, value=val)
        ws.column_dimensions[col_letter].hidden = True
        dv = DataValidation(
            type="list",
            formula1=f"=${col_letter}$1:${col_letter}${len(values)}",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
        )
        ws.add_data_validation(dv)
        tgt_letter = get_column_letter(target_col)
        dv.add(f"{tgt_letter}{data_start_row}:{tgt_letter}{data_end_row}")
        list_col += 1

    # category level validations
    for depth in range(0, max_depth + 1):
        values = levels.get(depth, []) or []
        _add_validation(depth + 1, values)

    # attribute dict validations
    for idx_attr, a in enumerate(attr_defs):
        options = a.get("options") or {}
        dict_id = str(options.get("dict_id") or "").strip()
        if not dict_id:
            continue
        dict_values: List[str] = []
        if dict_id:
            from app.storage.json_store import load_dict, dict_exists

            def _load_values(did: str) -> List[str]:
                doc = load_dict(did)
                items = doc.get("items") or []
                out_vals = []
                for it in items:
                    if isinstance(it, dict):
                        out_vals.append(str(it.get("value") or "").strip())
                    else:
                        out_vals.append(str(it).strip())
                return [v for v in out_vals if v]

            dict_values = _load_values(dict_id)
            if not dict_values and dict_id and not dict_id.startswith("dict_"):
                alt_id = f"dict_{dict_id}"
                if dict_exists(alt_id):
                    dict_values = _load_values(alt_id)

        if dict_values:
            target_col = base_start + 3 + idx_attr
            _add_validation(target_col, sorted(set(dict_values)))

    import io

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"products_{category_id}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/catalog/products/import.xlsx")
async def import_products_xlsx(
    file: UploadFile = File(...),
    category_id: Optional[str] = Query(default=None),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload .xlsx file")

    from openpyxl import load_workbook
    import io

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": True, "created": 0, "items": []}

    def _row_norm(row):
        return [str(x or "").strip() for x in (row or [])]

    header1 = _row_norm(rows[0] if rows else [])
    header2 = _row_norm(rows[1] if len(rows) > 1 else [])

    def _looks_like_code_row(row: List[str]) -> bool:
        row_l = [r.lower() for r in row if r]
        keys = {"title", "sku_pim", "sku_gt", "category_1", "category_id"}
        return any(k in row_l for k in keys)

    if header2 and _looks_like_code_row(header2):
        header = header2
        data_rows = rows[2:]
    else:
        header = header1
        data_rows = rows[1:]

    header_l = [h.lower() for h in header]
    idx_title = (
        header_l.index("title")
        if "title" in header_l
        else header_l.index("название")
        if "название" in header_l
        else -1
    )
    idx_cat = header_l.index("category_id") if "category_id" in header_l else -1
    idx_sku_pim = header_l.index("sku_pim") if "sku_pim" in header_l else -1
    idx_sku_gt = header_l.index("sku_gt") if "sku_gt" in header_l else -1

    cat_cols = [i for i, h in enumerate(header_l) if h.startswith("category_")]

    if idx_title < 0:
        raise HTTPException(status_code=400, detail="title column is required")

    nodes = _load_nodes()
    by_id = {n.get("id"): n for n in nodes}
    path_cache: Dict[str, List[str]] = {}

    def _path_names(node_id: str) -> List[str]:
        if node_id in path_cache:
            return path_cache[node_id]
        node = by_id.get(node_id)
        if not node:
            return []
        parent_id = node.get("parent_id")
        prefix = _path_names(parent_id) if parent_id else []
        names = prefix + [str(node.get("name") or "").strip()]
        path_cache[node_id] = names
        return names

    path_map: Dict[Tuple[str, ...], str] = {}
    for nid in by_id.keys():
        names = [n for n in _path_names(nid) if n]
        if not names:
            continue
        path_map[tuple([n.lower() for n in names])] = nid

    prepared: List[Dict[str, Any]] = []
    for r in data_rows:
        if not r:
            continue
        title = str((r[idx_title] if idx_title < len(r) else "") or "").strip()
        if not title:
            continue
        cid = ""
        if cat_cols:
            names = []
            for i in cat_cols:
                if i >= len(r):
                    continue
                val = str(r[i] or "").strip()
                if val:
                    names.append(val)
            key = tuple([n.lower() for n in names if n])
            if key:
                cid = path_map.get(key, "")
                if not cid:
                    raise HTTPException(status_code=400, detail=f"category_path not found: {' / '.join(names)}")

        if not cid and idx_cat >= 0:
            cid = str((r[idx_cat] if idx_cat < len(r) else "") or "").strip()
        if not cid:
            cid = str(category_id or "").strip()
        if not cid:
            raise HTTPException(status_code=400, detail="category_id is required")
        sku_pim = str((r[idx_sku_pim] if idx_sku_pim >= 0 and idx_sku_pim < len(r) else "") or "").strip()
        sku_gt = str((r[idx_sku_gt] if idx_sku_gt >= 0 and idx_sku_gt < len(r) else "") or "").strip()
        prepared.append(
            {
                "title": title,
                "category_id": cid,
                "sku_pim": sku_pim,
                "sku_gt": sku_gt,
            }
        )

    if not prepared:
        return {"ok": True, "created": 0, "items": []}

    auto_needed = [p for p in prepared if not (p.get("sku_pim") and p.get("sku_gt"))]
    sku_items = allocate_sku_pairs_service(len(auto_needed)).get("items") or []
    sku_iter = iter(sku_items)
    created_items = []
    for row in prepared:
        sku = {}
        if not (row.get("sku_pim") and row.get("sku_gt")):
            sku = next(sku_iter, {})
        payload = {
            "category_id": row["category_id"],
            "type": "single",
            "title": row["title"],
            "sku_pim": row.get("sku_pim") or sku.get("sku_pim"),
            "sku_gt": row.get("sku_gt") or sku.get("sku_gt"),
            "selected_params": [],
            "feature_params": [],
            "exports_enabled": {},
        }
        p = create_product_service(payload)
        created_items.append({"id": p.get("id"), "title": p.get("title"), "category_id": p.get("category_id")})

    return {"ok": True, "created": len(created_items), "items": created_items}


@router.post("/catalog/products")
def create_product(payload: ProductCreate):
    nodes = _load_nodes()
    by_id = {n["id"]: n for n in nodes}
    if payload.category_id not in by_id:
        raise HTTPException(status_code=404, detail="Category not found")

    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Empty product name")

    products = _load_products()

    product_id = str(uuid.uuid4())
    product = {
        "id": product_id,
        "name": name,
        "category_id": payload.category_id,
    }
    products.append(product)
    _save_products(products)

    return {"id": product_id}


@router.post("/catalog/products/bulk-delete")
def bulk_delete_products(payload: ProductDeleteReq):
    ids = [str(x).strip() for x in (payload.ids or []) if str(x).strip()]
    if not ids:
        raise HTTPException(status_code=400, detail="ids required")
    result = delete_products_bulk_service(ids)
    return result
